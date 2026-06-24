"""
Retail Analytics  Excel Dashboard Generator
Connects to MySQL, runs all analytics queries, and produces a
fully formatted Excel workbook with charts and KPI cards.
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import mysql.connector
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import warnings
warnings.filterwarnings("ignore")

#  DB CONFIG 
DB_CONFIG = {
    "host":     "localhost",
    "user":     "root",
    "password": "Admin@123",   
    "database": "retail_analytics"
}

#  COLOUR PALETTE 
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "BDD7EE"
ACCENT_GOLD = "FFD700"
WHITE       = "FFFFFF"
LIGHT_GREY  = "F2F2F2"
GREEN       = "70AD47"
RED         = "FF0000"
DARK_GREY   = "404040"

#  HELPERS 
def connect():
    return mysql.connector.connect(**DB_CONFIG)

def run_query(sql: str) -> pd.DataFrame:
    conn = connect()
    df = pd.read_sql(sql, conn)
    conn.close()
    return df

def header_style(ws, cell_ref, value, bg=DARK_BLUE, fg=WHITE, size=11, bold=True):
    cell = ws[cell_ref]
    cell.value = value
    cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def kpi_card(ws, row, col, label, value, bg=MID_BLUE):
    lc = ws.cell(row=row,   column=col, value=label)
    vc = ws.cell(row=row+1, column=col, value=value)
    for c in (lc, vc):
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(name="Arial", bold=True, color=WHITE, size=12 if c == lc else 16)
    ws.row_dimensions[row].height   = 22
    ws.row_dimensions[row+1].height = 30

def style_table(ws, start_row, end_row, start_col, end_col):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(start_row, end_row+1):
        for c in range(start_col, end_col+1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.font = Font(name="Arial", size=10)
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)
            cell.alignment = Alignment(vertical="center")

def write_df(ws, df, start_row=1, start_col=1, header_bg=MID_BLUE):
    for ci, col_name in enumerate(df.columns, start=start_col):
        header_style(ws, ws.cell(row=start_row, column=ci).coordinate,
                     str(col_name).replace("_", " ").title(), bg=header_bg)
    for ri, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row+1):
        for ci, val in enumerate(row_data, start=start_col):
            ws.cell(row=ri, column=ci, value=val)
    style_table(ws, start_row+1, start_row+len(df), start_col, start_col+len(df.columns)-1)
    for ci, col in enumerate(df.columns, start=start_col):
        max_len = max(df[col].astype(str).map(len).max(), len(col)) + 4
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len, 30)

#  QUERIES 
SQL = {
    "monthly": """
        SELECT DATE_FORMAT(o.order_date,'%Y-%m') AS month,
               COUNT(DISTINCT o.order_id)         AS total_orders,
               COUNT(DISTINCT o.customer_id)      AS unique_customers,
               ROUND(SUM(oi.line_total),2)         AS gross_revenue,
               ROUND(SUM(oi.unit_cost*oi.quantity),2) AS total_cogs,
               ROUND(SUM(oi.line_total - oi.unit_cost*oi.quantity),2) AS gross_profit,
               ROUND(SUM(oi.line_total - oi.unit_cost*oi.quantity)
                     /NULLIF(SUM(oi.line_total),0)*100,2) AS margin_pct
        FROM fact_orders o JOIN fact_order_items oi ON o.order_id=oi.order_id
        WHERE o.status NOT IN ('Cancelled','Returned')
        GROUP BY DATE_FORMAT(o.order_date,'%Y-%m') ORDER BY month
    """,
    "top_products": """
        SELECT p.product_name, c.category_name,
               SUM(oi.quantity) AS units_sold,
               ROUND(SUM(oi.line_total),2) AS revenue,
               ROUND(SUM(oi.line_total-oi.unit_cost*oi.quantity),2) AS profit,
               ROUND(SUM(oi.line_total-oi.unit_cost*oi.quantity)
                     /NULLIF(SUM(oi.line_total),0)*100,2) AS margin_pct
        FROM fact_order_items oi
        JOIN dim_products p   ON oi.product_id=p.product_id
        JOIN dim_categories c ON p.category_id=c.category_id
        JOIN fact_orders o    ON oi.order_id=o.order_id
        WHERE o.status NOT IN ('Cancelled','Returned')
        GROUP BY p.product_id,p.product_name,c.category_name
        ORDER BY revenue DESC LIMIT 10
    """,
    "region": """
        SELECT r.region_name,
               COUNT(DISTINCT o.order_id)    AS total_orders,
               ROUND(SUM(oi.line_total),2)   AS revenue,
               ROUND(AVG(o.total_amount),2)  AS avg_order_value,
               COUNT(DISTINCT o.customer_id) AS unique_customers
        FROM dim_regions r
        JOIN fact_orders o       ON r.region_id=o.region_id
        JOIN fact_order_items oi ON o.order_id=oi.order_id
        WHERE o.status NOT IN ('Cancelled','Returned')
        GROUP BY r.region_id,r.region_name ORDER BY revenue DESC
    """,
    "sales_rep": """
        SELECT sr.full_name AS sales_rep, r.region_name,
               COUNT(DISTINCT o.order_id)   AS deals_closed,
               ROUND(SUM(oi.line_total),2)  AS revenue,
               ROUND(SUM(oi.line_total-oi.unit_cost*oi.quantity),2) AS gross_profit
        FROM dim_salesreps sr
        JOIN fact_orders o       ON sr.rep_id=o.rep_id
        JOIN fact_order_items oi ON o.order_id=oi.order_id
        JOIN dim_regions r       ON sr.region_id=r.region_id
        WHERE o.status NOT IN ('Cancelled','Returned')
        GROUP BY sr.rep_id,sr.full_name,r.region_name ORDER BY revenue DESC
    """,
    "category": """
        SELECT COALESCE(parent.category_name,c.category_name) AS parent_category,
               c.category_name,
               SUM(oi.quantity) AS units_sold,
               ROUND(SUM(oi.line_total),2) AS revenue,
               ROUND(SUM(oi.line_total-oi.unit_cost*oi.quantity),2) AS profit
        FROM fact_order_items oi
        JOIN dim_products p   ON oi.product_id=p.product_id
        JOIN dim_categories c ON p.category_id=c.category_id
        LEFT JOIN dim_categories parent ON c.parent_id=parent.category_id
        JOIN fact_orders o    ON oi.order_id=o.order_id
        WHERE o.status NOT IN ('Cancelled','Returned')
        GROUP BY parent_category,c.category_id,c.category_name ORDER BY revenue DESC
    """,
    "clv": """
        SELECT c.full_name, c.customer_segment, r.region_name,
               COUNT(DISTINCT o.order_id)    AS total_orders,
               ROUND(SUM(oi.line_total),2)   AS lifetime_value,
               ROUND(AVG(o.total_amount),2)  AS avg_order_value,
               MIN(DATE(o.order_date))        AS first_order,
               MAX(DATE(o.order_date))        AS last_order
        FROM dim_customers c
        JOIN fact_orders o       ON c.customer_id=o.customer_id
        JOIN fact_order_items oi ON o.order_id=oi.order_id
        JOIN dim_regions r       ON c.region_id=r.region_id
        WHERE o.status NOT IN ('Cancelled','Returned')
        GROUP BY c.customer_id,c.full_name,c.customer_segment,r.region_name
        ORDER BY lifetime_value DESC LIMIT 20
    """,
    "payment": """
        SELECT payment_method,
               COUNT(*) AS order_count,
               ROUND(COUNT(*)*100.0/SUM(COUNT(*)) OVER(),2) AS pct_orders,
               ROUND(SUM(total_amount),2) AS revenue
        FROM fact_orders WHERE status NOT IN ('Cancelled','Returned')
        GROUP BY payment_method ORDER BY order_count DESC
    """,
    "yoy": """
        SELECT YEAR(o.order_date) AS year,
               ROUND(SUM(oi.line_total),2) AS revenue,
               COUNT(DISTINCT o.order_id)  AS orders,
               ROUND(SUM(oi.line_total-oi.unit_cost*oi.quantity),2) AS profit
        FROM fact_orders o JOIN fact_order_items oi ON o.order_id=oi.order_id
        WHERE o.status NOT IN ('Cancelled','Returned')
        GROUP BY YEAR(o.order_date) ORDER BY year
    """
}

#  SHEET BUILDERS 

def build_cover(wb, data):
    ws = wb.create_sheet(" Dashboard")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for col in "BCDEFGHIJ":
        ws.column_dimensions[col].width = 18

    # Title banner
    ws.merge_cells("B1:J3")
    title_cell = ws["B1"]
    title_cell.value = "  RETAIL ANALYTICS DASHBOARD"
    title_cell.font = Font(name="Arial", bold=True, size=22, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 25

    # Sub-title
    ws.merge_cells("B4:J4")
    sub = ws["B4"]
    sub.value = f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}  |  Data Source: MySQL retail_analytics"
    sub.font = Font(name="Arial", italic=True, size=10, color=DARK_GREY)
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[4].height = 18

    # KPI Row
    ws.row_dimensions[6].height = 22
    ws.row_dimensions[7].height = 32
    monthly = data["monthly"]
    total_rev   = monthly["gross_revenue"].sum()
    total_profit= monthly["gross_profit"].sum()
    total_orders= monthly["total_orders"].sum()
    avg_margin  = round(total_profit / total_rev * 100, 1) if total_rev else 0
    top_region  = data["region"].iloc[0]["region_name"] if not data["region"].empty else "N/A"

    kpis = [
        (" Total Revenue",    f"{total_rev:,.0f}",    MID_BLUE),
        (" Total Orders",     f"{total_orders:,}",      "2E7D32"),
        (" Gross Profit",     f"{total_profit:,.0f}", "6A1B9A"),
        (" Avg Margin",       f"{avg_margin}%",         "E65100"),
        (" Top Region",       top_region,               DARK_BLUE),
    ]
    for i, (label, value, color) in enumerate(kpis, start=2):
        ws.merge_cells(start_row=6, start_column=i, end_row=6, end_column=i)
        ws.merge_cells(start_row=7, start_column=i, end_row=7, end_column=i)
        kpi_card(ws, 6, i, label, value, bg=color)

    # Mini monthly revenue table
    ws.merge_cells("B9:J9")
    hdr = ws["B9"]
    hdr.value = "Monthly Revenue at a Glance"
    hdr.font = Font(name="Arial", bold=True, size=12, color=WHITE)
    hdr.fill = PatternFill("solid", fgColor=MID_BLUE)
    hdr.alignment = Alignment(horizontal="center")

    cols = ["month","total_orders","gross_revenue","gross_profit","margin_pct"]
    mini = data["monthly"][cols].tail(12)
    write_df(ws, mini, start_row=10, start_col=2, header_bg=MID_BLUE)

    # Revenue bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Monthly Revenue vs Profit"
    chart.style = 10
    chart.y_axis.title = "Amount ()"
    chart.x_axis.title = "Month"
    chart.shape = 4
    n = len(mini)
    data_ref  = Reference(ws, min_col=4, max_col=5, min_row=10, max_row=10+n)
    cats_ref  = Reference(ws, min_col=2, min_row=11, max_row=10+n)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width  = 22
    chart.height = 12
    ws.add_chart(chart, "B23")
    return ws


def build_sheet(wb, title, df, chart_type=None, chart_col=None, chart_val_col=None):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    ws.merge_cells(f"B1:{get_column_letter(len(df.columns)+1)}1")
    hdr = ws["B1"]
    hdr.value = title.replace("","").replace("","").replace("","").replace("","").strip()
    hdr.font = Font(name="Arial", bold=True, size=14, color=WHITE)
    hdr.fill = PatternFill("solid", fgColor=DARK_BLUE)
    hdr.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    write_df(ws, df, start_row=3, start_col=2)

    if chart_type and chart_col and chart_val_col:
        n = len(df)
        if chart_type == "bar":
            chart = BarChart()
            chart.type = "bar"
            chart.style = 10
        elif chart_type == "pie":
            chart = PieChart()
            chart.style = 10
        else:
            chart = BarChart()

        val_ref  = Reference(ws, min_col=chart_val_col+1, min_row=3, max_row=3+n)
        cats_ref = Reference(ws, min_col=chart_col+1,     min_row=4, max_row=3+n)
        chart.add_data(val_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.width  = 20
        chart.height = 12
        chart.title  = title.split(" ", 1)[-1]
        anchor_col   = get_column_letter(len(df.columns)+3)
        ws.add_chart(chart, f"{anchor_col}3")
    return ws


#  MAIN 
def main():
    print(" Fetching data from MySQL...")
    data = {k: run_query(v) for k, v in SQL.items()}
    print(" All queries complete.")

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    print(" Building Dashboard sheet...")
    build_cover(wb, data)

    print(" Building Monthly Trends sheet...")
    build_sheet(wb, " Monthly Trends", data["monthly"],
                chart_type="bar", chart_col=0, chart_val_col=3)

    print(" Building Top Products sheet...")
    build_sheet(wb, " Top Products", data["top_products"],
                chart_type="bar", chart_col=0, chart_val_col=3)

    print(" Building Regional Sales sheet...")
    build_sheet(wb, " Regional Sales", data["region"],
                chart_type="bar", chart_col=0, chart_val_col=2)

    print(" Building Sales Reps sheet...")
    build_sheet(wb, " Sales Reps", data["sales_rep"],
                chart_type="bar", chart_col=0, chart_val_col=3)

    print(" Building Category Analysis sheet...")
    build_sheet(wb, " Categories", data["category"],
                chart_type="bar", chart_col=1, chart_val_col=3)

    print(" Building CLV sheet...")
    build_sheet(wb, " Customer CLV", data["clv"])

    print(" Building Payment Methods sheet...")
    build_sheet(wb, " Payments", data["payment"],
                chart_type="pie", chart_col=0, chart_val_col=1)

    print(" Building YoY Growth sheet...")
    build_sheet(wb, " YoY Growth", data["yoy"],
                chart_type="bar", chart_col=0, chart_val_col=1)

    out = "retail_analytics_dashboard.xlsx"
    wb.save(out)
    print(f"\n Dashboard saved  {out}")

if __name__ == "__main__":
    main()
